import csv
import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import permissions, status
from rest_framework.response import Response
from django.utils import timezone

from accounts.models import ResidentProfile, VolunteerProfile, SecurityProfile, GuardianProfile, UserDocument
from society.models import Society, BlockTower, Flat

from sos.models import SOSIncident, EmergencyCategory

class ReportDownloadAPIView(APIView):
    """
    GET /api/reports/download/?type=residents|volunteers|security|verifications|societies|incidents&format=csv|excel|pdf
    Generates dynamic downloadable reports for gated community management and emergency analytics.
    Supports global filters: date_from, date_to, timeframe, society, category, priority, status.
    """
    permission_classes = [permissions.IsAuthenticated]

    def perform_content_negotiation(self, request, force=False):
        # Override DRF default content negotiation so query parameter `format=pdf|excel` does not trigger Http404
        renderers_list = self.get_renderers()
        return (renderers_list[0], renderers_list[0].media_type)

    def get(self, request):
        print(f"ReportDownloadAPIView called by user={request.user} role={request.user.role}")
        # Admin authorization verification for platform-wide reports
        if request.user.role not in ["ADMIN", "SUPERUSER", "STAFF"] and not request.user.is_staff:
            return Response({"detail": "Only authorized Admin personnel can export platform-wide reports."}, status=status.HTTP_403_FORBIDDEN)

        report_type = request.query_params.get("type", "residents").lower()
        export_format = request.query_params.get("format", "csv").lower()

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"CareConnect_{report_type.title()}_Report_{timestamp}"

        if report_type in ["incidents", "sos", "emergencies"]:
            headers = ["ID", "Created At", "Resident Name", "Society", "Emergency Category", "Priority", "Status", "Response Time", "Assigned Responder"]
            incidents_qs = SOSIncident.objects.select_related("resident", "category", "assigned_responder", "resident__resident_profile__society").all()

            # Global Filters
            society_param = request.query_params.get("society") or request.query_params.get("society_id")
            if society_param and str(society_param).strip() != "" and str(society_param).lower() != "all":
                incidents_qs = incidents_qs.filter(resident__resident_profile__society_id=society_param)

            category_param = request.query_params.get("category") or request.query_params.get("category_id")
            if category_param and str(category_param).strip() != "" and str(category_param).lower() != "all":
                if str(category_param).isdigit():
                    incidents_qs = incidents_qs.filter(category_id=int(category_param))
                else:
                    incidents_qs = incidents_qs.filter(category__name__icontains=category_param)

            priority_param = request.query_params.get("priority")
            if priority_param and str(priority_param).strip() != "" and str(priority_param).lower() != "all":
                incidents_qs = incidents_qs.filter(priority__iexact=priority_param)

            status_param = request.query_params.get("status")
            if status_param and str(status_param).strip() != "" and str(status_param).lower() != "all":
                incidents_qs = incidents_qs.filter(status__iexact=status_param)

            # Date Range / Timeframe Filter
            timeframe = request.query_params.get("timeframe")
            now = timezone.now()
            if timeframe == "today":
                incidents_qs = incidents_qs.filter(created_at__gte=now.replace(hour=0, minute=0, second=0))
            elif timeframe == "yesterday":
                yest = now - timezone.timedelta(days=1)
                incidents_qs = incidents_qs.filter(created_at__range=(yest.replace(hour=0, minute=0, second=0), yest.replace(hour=23, minute=59, second=59)))
            elif timeframe == "7days":
                incidents_qs = incidents_qs.filter(created_at__gte=now - timezone.timedelta(days=7))
            elif timeframe == "30days":
                incidents_qs = incidents_qs.filter(created_at__gte=now - timezone.timedelta(days=30))
            elif timeframe == "90days":
                incidents_qs = incidents_qs.filter(created_at__gte=now - timezone.timedelta(days=90))
            elif timeframe == "this_year":
                incidents_qs = incidents_qs.filter(created_at__year=now.year)

            date_from = request.query_params.get("date_from")
            if date_from:
                incidents_qs = incidents_qs.filter(created_at__gte=date_from)

            date_to = request.query_params.get("date_to")
            if date_to:
                incidents_qs = incidents_qs.filter(created_at__lte=date_to)

            rows = []
            for inc in incidents_qs.order_by("-created_at"):
                resp_str = "N/A"
                if inc.accepted_at and inc.accepted_at >= inc.created_at:
                    sec = round((inc.accepted_at - inc.created_at).total_seconds())
                    resp_str = f"{sec // 60}m {sec % 60}s"

                soc_name = "N/A"
                try:
                    if inc.resident and inc.resident.resident_profile and inc.resident.resident_profile.society:
                        soc_name = inc.resident.resident_profile.society.name
                except Exception:
                    soc_name = "N/A"

                rows.append([
                    f"#{inc.id}",
                    inc.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    inc.resident.full_name if inc.resident else "Unknown",
                    soc_name,
                    inc.category.name if inc.category else "Emergency",
                    inc.priority or "NORMAL",
                    inc.current_status,
                    resp_str,
                    inc.assigned_responder.full_name if inc.assigned_responder else "Unassigned"
                ])

        elif report_type == "residents":
            headers = ["ID", "Resident Name", "Email", "Phone", "Society", "Block", "Flat", "Status", "Verification Date", "Remarks"]
            rows = []
            for rp in ResidentProfile.objects.select_related("user", "society", "block", "flat").all():
                rows.append([
                    rp.id,
                    rp.user.full_name,
                    rp.user.email,
                    rp.user.phone_number,
                    rp.society.name if rp.society else "Unassigned",
                    rp.block.name if rp.block else "Unassigned",
                    rp.flat.flat_number if rp.flat else "Unassigned",
                    rp.status,
                    rp.verification_date.strftime("%Y-%m-%d %H:%M") if rp.verification_date else "N/A",
                    rp.remarks or ""
                ])

        elif report_type == "volunteers":
            headers = ["ID", "Volunteer Name", "Email", "Phone", "Assigned Society", "Assigned Block", "Skills", "Emergency Training", "Status", "Online State"]
            rows = []
            for vp in VolunteerProfile.objects.select_related("user", "assigned_society", "assigned_block").all():
                rows.append([
                    vp.id,
                    vp.user.full_name,
                    vp.user.email,
                    vp.user.phone_number,
                    vp.assigned_society.name if vp.assigned_society else "All Societies",
                    vp.assigned_block.name if vp.assigned_block else "All Blocks",
                    vp.skills,
                    vp.emergency_training or "General",
                    vp.status,
                    "Online" if vp.is_online else "Offline"
                ])

        elif report_type == "security":
            headers = ["ID", "Employee ID", "Security Name", "Email", "Phone", "Shift", "Assigned Society", "Assigned Block", "Verification Status", "Employment Status"]
            rows = []
            for sp in SecurityProfile.objects.select_related("user", "assigned_society", "assigned_block").all():
                rows.append([
                    sp.id,
                    sp.employee_id or f"SEC-{sp.id:04d}",
                    sp.user.full_name,
                    sp.user.email,
                    sp.user.phone_number,
                    sp.shift or "General",
                    sp.assigned_society.name if sp.assigned_society else "Unassigned",
                    sp.assigned_block.name if sp.assigned_block else "All Blocks",
                    sp.verification_status,
                    sp.employment_status
                ])

        elif report_type == "verifications":
            headers = ["ID", "User Name", "Email", "Role", "Verification Category", "Status", "Verification Date", "Verified By", "Remarks"]
            rows = []
            for rp in ResidentProfile.objects.select_related("user", "verified_by").all():
                rows.append([
                    f"RES-{rp.id}", rp.user.full_name, rp.user.email, rp.user.role, "Resident",
                    rp.status, rp.verification_date.strftime("%Y-%m-%d") if rp.verification_date else "N/A",
                    rp.verified_by.full_name if rp.verified_by else "Admin", rp.remarks or ""
                ])
            for vp in VolunteerProfile.objects.select_related("user", "verified_by").all():
                rows.append([
                    f"VOL-{vp.id}", vp.user.full_name, vp.user.email, vp.user.role, "Volunteer",
                    vp.status, vp.verification_date.strftime("%Y-%m-%d") if vp.verification_date else "N/A",
                    vp.verified_by.full_name if vp.verified_by else "Admin", vp.remarks or ""
                ])
            for sp in SecurityProfile.objects.select_related("user", "verified_by").all():
                rows.append([
                    f"SEC-{sp.id}", sp.user.full_name, sp.user.email, sp.user.role, "Security",
                    sp.verification_status, sp.verification_date.strftime("%Y-%m-%d") if sp.verification_date else "N/A",
                    sp.verified_by.full_name if sp.verified_by else "Admin", sp.remarks or ""
                ])

        elif report_type == "societies":
            headers = ["ID", "Society Code", "Society Name", "City", "State", "Pincode", "Manager", "Total Blocks", "Total Flats", "Status", "Created Date"]
            rows = []
            for s in Society.objects.select_related("society_manager").all():
                rows.append([
                    s.id,
                    s.code or f"SOC-{s.id:03d}",
                    s.name,
                    s.city,
                    s.state,
                    s.pincode,
                    s.society_manager.full_name if s.society_manager else (s.contact_person or "Unassigned"),
                    s.blocks.count(),
                    Flat.objects.filter(block__society=s).count(),
                    s.status,
                    s.created_at.strftime("%Y-%m-%d")
                ])
        else:
            return Response({"detail": "Invalid report type specified."}, status=status.HTTP_400_BAD_REQUEST)

        # Generate Output Formats
        if export_format in ["excel", "xlsx"]:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"{report_type.title()} Report"

                # Title Banner
                ws.merge_cells("A1:I1")
                title_cell = ws["A1"]
                title_cell.value = f"CareConnect Platform — {report_type.title()} Official Report"
                title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="E93F41", end_color="E93F41", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 30

                # Meta Row
                ws.merge_cells("A2:I2")
                meta_cell = ws["A2"]
                meta_cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')} | Total Records: {len(rows)}"
                meta_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
                meta_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 20

                # Header Row
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0B1220", end_color="0B1220", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin', color='CBD5E1'),
                    right=Side(style='thin', color='CBD5E1'),
                    top=Side(style='thin', color='CBD5E1'),
                    bottom=Side(style='thin', color='CBD5E1')
                )

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws.row_dimensions[4].height = 24

                # Data Rows
                for row_idx, row_data in enumerate(rows, 5):
                    for col_idx, cell_val in enumerate(row_data, 1):
                        c = ws.cell(row=row_idx, column=col_idx, value=str(cell_val))
                        c.font = Font(name="Calibri", size=10)
                        c.border = thin_border
                        if col_idx in [1, 6, 7, 8]:
                            c.alignment = Alignment(horizontal="center")

                # Auto-fit columns
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(
                    output.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
                return response
            except ImportError:
                # Fallback to CSV if openpyxl unavailable
                pass

        if export_format in ["csv"]:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'

            writer = csv.writer(response)
            writer.writerow([f"CareConnect Platform — {report_type.title()} Report"])
            writer.writerow([f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"])
            writer.writerow([])
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            return response

        elif export_format == "pdf":
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'

            buffer = io.BytesIO()
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors

                pdf_doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
                elements = []
                styles = getSampleStyleSheet()

                title_style = ParagraphStyle(
                    'TitleStyle',
                    parent=styles['Title'],
                    fontName='Helvetica-Bold',
                    fontSize=16,
                    textColor=colors.HexColor('#E93F41'),
                    alignment=0,
                    spaceAfter=4
                )
                subtitle_style = ParagraphStyle(
                    'SubtitleStyle',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=9,
                    textColor=colors.HexColor('#64748B'),
                    spaceAfter=12
                )

                elements.append(Paragraph(f"<b>CareConnect Emergency Response Platform</b>", title_style))
                elements.append(Paragraph(f"<b>Report:</b> {report_type.title()} Audit | <b>Generated At:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')} | <b>Total Records:</b> {len(rows)}", subtitle_style))
                elements.append(Spacer(1, 10))

                table_data = [headers] + [[str(cell) for cell in row] for row in rows]
                t = Table(table_data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E93F41')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ]))
                elements.append(t)
                pdf_doc.build(elements)
                pdf_bytes = buffer.getvalue()
                buffer.close()
            except ImportError:
                content = f"CARECONNECT PLATFORM REPORT\nType: {report_type.upper()}\nGenerated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                content += "\t".join(headers) + "\n"
                for row in rows:
                    content += "\t".join([str(v) for v in row]) + "\n"
                pdf_bytes = content.encode('utf-8')

            response.write(pdf_bytes)
            return response

        return Response({"detail": "Invalid export format specified."}, status=status.HTTP_400_BAD_REQUEST)
